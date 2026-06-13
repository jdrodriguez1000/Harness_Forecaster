"""Generador reproducible de los fixtures del PASO 12 (E9).

Materializa los ~20 archivos de prueba en este mismo directorio con bytes y
encoding exactos. Los binarios Excel se generan con openpyxl (.xlsx) y xlwt (.xls
— solo build-time; el runtime del 015 lee .xls con xlrd).

Uso:  python tests/fixtures/_build_fixtures.py
La expectativa de cada fixture está documentada en tests/fixtures/README.md.
"""

from __future__ import annotations

import os

HERE = os.path.dirname(os.path.abspath(__file__))


def _w(nombre: str, data: bytes) -> None:
    with open(os.path.join(HERE, nombre), "wb") as fh:
        fh.write(data)


def _csv(filas: list[str], sep: str, encoding: str, bom: bool = False) -> bytes:
    texto = "\n".join(sep.join(f) for f in filas) + "\n"
    if bom and encoding == "utf-8":
        return texto.encode("utf-8-sig")
    return texto.encode(encoding)


# --- CSV / texto -----------------------------------------------------------

_OK = [
    ["fecha_pedido", "id_cliente", "id_producto", "cantidad_solicitada"],
    ["2024-01-01", "C1", "P1", "5"],
    ["2024-06-01", "C2", "P2", "3"],
    ["2025-01-01", "C3", "P3", "8"],
]


def build_csv():
    _w("orders_comma_utf8.csv", _csv(_OK, ",", "utf-8"))
    _w("orders_semicolon.csv", _csv(_OK, ";", "utf-8"))
    _w("orders_pipe.csv", _csv(_OK, "|", "utf-8"))
    _w("orders_utf8sig_bom.csv", _csv(_OK, ",", "utf-8", bom=True))

    # cp1252 con acentos (EL CANARIO): í=0xED, ü/ó presentes
    cp = [
        ["fecha_pedido", "id_cliente", "id_producto", "cantidad_solicitada"],
        ["2024-01-01", "Pingüino", "Camión", "5"],
        ["2024-06-01", "Categoría", "Limón", "3"],
    ]
    _w("cp1252_acentos.csv", _csv(cp, ";", "cp1252"))
    _w("bronce_bitexacto.csv", _csv(cp, ";", "cp1252"))

    # Estructura: falta id_producto -> rechazo (GATE P2)
    falta = [
        ["fecha_pedido", "id_cliente", "cantidad_solicitada"],
        ["2024-01-01", "C1", "5"],
    ]
    _w("missing_id_producto.csv", _csv(falta, ";", "utf-8"))

    # Estructura: 4 mínimos + 3 ideales + 1 columna no-ideal -> aprobado + ideales
    extra = [
        ["fecha_pedido", "id_cliente", "id_producto", "cantidad_solicitada",
         "fecha_entrega_solicitada", "estado_pedido", "ciudad", "columna_extra"],
        ["2024-01-01", "C1", "P1", "5", "2024-01-05", "entregado", "Bogotá", "x"],
        ["2024-06-01", "C2", "P2", "3", "2024-06-04", "pendiente", "Medellín", "y"],
    ]
    _w("extra_columns.csv", _csv(extra, ";", "cp1252"))

    # Tipos: 3 negativas + 2 no-numéricas = 5 errores en cantidad (no elimina)
    neg = [
        ["fecha_pedido", "id_cliente", "id_producto", "cantidad_solicitada"],
        ["2024-01-01", "C1", "P1", "-1"],
        ["2024-01-02", "C2", "P2", "-2"],
        ["2024-01-03", "C3", "P3", "-3"],
        ["2024-01-04", "C4", "P4", "abc"],
        ["2024-01-05", "C5", "P5", "xyz"],
        ["2024-01-06", "C6", "P6", "10"],
    ]
    _w("cantidad_negativa.csv", _csv(neg, ";", "utf-8"))

    # Tipos: 3 formatos de fecha distintos, todos parseables -> 0 errores fecha
    fechas = [
        ["fecha_pedido", "id_cliente", "id_producto", "cantidad_solicitada"],
        ["2024-01-01", "C1", "P1", "1"],
        ["15/06/2024", "C2", "P2", "2"],
        ["20-03-2024", "C3", "P3", "3"],
    ]
    _w("fechas_3_formatos.csv", _csv(fechas, ";", "utf-8"))

    # Tipos: 2 id_cliente vacíos -> 2 errores texto
    vacio_cli = [
        ["fecha_pedido", "id_cliente", "id_producto", "cantidad_solicitada"],
        ["2024-01-01", "", "P1", "1"],
        ["2024-01-02", "C2", "P2", "2"],
        ["2024-01-03", "   ", "P3", "3"],
    ]
    _w("id_cliente_vacio.csv", _csv(vacio_cli, ";", "utf-8"))

    # Duplicados internos (batch): 4 filas repetidas exactas en la clave compuesta
    dup = [["fecha_pedido", "id_cliente", "id_producto", "cantidad_solicitada"]]
    for i in range(6):
        dup.append(["2024-01-01", f"C{i}", "P1", "1"])
    for _ in range(4):
        dup.append(["2024-01-01", "C0", "P1", "1"])  # 4 duplicados de la 1ª clave
    _w("dup_internos_batch.csv", _csv(dup, ";", "utf-8"))

    # Incremental: 6 claves C1..C6 (el test arma un Bronce previo con C1..C3)
    inc = [["fecha_pedido", "id_cliente", "id_producto", "cantidad_solicitada"]]
    for i in range(1, 7):
        inc.append(["2024-01-01", f"C{i}", "P1", "1"])
    _w("incremental_repetidos.csv", _csv(inc, ";", "utf-8"))

    # Rango: ~3.4 años reales (2021-01-01 → 2024-06-01); declarado 2 -> warning
    rango = [
        ["fecha_pedido", "id_cliente", "id_producto", "cantidad_solicitada"],
        ["2021-01-01", "C1", "P1", "1"],
        ["2022-06-15", "C2", "P2", "2"],
        ["2024-06-01", "C3", "P3", "3"],
    ]
    _w("historial_34_anios.csv", _csv(rango, ";", "utf-8"))

    # Corrupto / vacío
    _w("vacio.csv", b"")
    _w("binario_corrupto.bin", bytes([0x00, 0xFF, 0x13, 0x37, 0x00, 0x90, 0x8D]) * 20)


# --- Excel -----------------------------------------------------------------

def build_xlsx():
    from openpyxl import Workbook

    # Cabecera en fila 1, hoja única
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Datos")
    for fila in _OK:
        ws.append(fila)
    wb.save(os.path.join(HERE, "header_row_1.xlsx"))

    # Cabecera en fila 3 (dos filas vacías arriba), hoja única
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Hoja1")
    ws.append([None, None, None, None])
    ws.append([None, None, None, None])
    for fila in _OK:
        ws.append(fila)
    wb.save(os.path.join(HERE, "header_row_3.xlsx"))

    # Multi-hoja -> ambiguo
    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("Ventas")
    for fila in _OK:
        ws1.append(fila)
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["total"])
    ws2.append([99])
    wb.save(os.path.join(HERE, "multi_sheet.xlsx"))


def build_xls():
    import xlwt

    wb = xlwt.Workbook()
    ws = wb.add_sheet("Datos")
    for r, fila in enumerate(_OK):
        for c, val in enumerate(fila):
            ws.write(r, c, val)
    wb.save(os.path.join(HERE, "legacy.xls"))


if __name__ == "__main__":
    build_csv()
    build_xlsx()
    build_xls()
    archivos = sorted(f for f in os.listdir(HERE) if not f.startswith("_") and f != "README.md")
    print(f"{len(archivos)} fixtures generados:")
    for a in archivos:
        print("  ", a)
