"""P1 — Detección de formato. format_detector.py

detect_format(raw_bytes, filename, huella=None) -> FormatSpec

Resuelve, sin alterar los bytes, qué hay realmente dentro del snapshot:
  - tipo: 'csv' | 'xlsx' | 'xls'  (magic bytes + extensión)
  - encoding (CSV): cascada estricta utf-8 → utf-8-sig → cp1252 → latin-1
  - delimitador (CSV): , ; |  por consistencia de columnas entre filas
  - hoja + fila_cabecera (Excel): heurística (o la huella de client_config)

Política (brief §2.4.2/2.4.3): si la heurística NO resuelve con confianza
→ ambiguo=True con la mejor propuesta → el processor escala al operador,
nunca adivina. Binario no tabular → corrupto=True (P1 rechaza).

EL CANARIO (LEC del plan): el acento cp1252 debe sobrevivir byte-a-byte. La
cascada decodifica en modo ESTRICTO (sin errors='replace'): el primer encoding
que decodifica sin excepción gana. utf-8 estricto falla ante 0xED suelto, así
que un archivo cp1252 con acentos cae correctamente a cp1252 (no a mojibake).

Ver plan/015_intake.md (PASO 5) y brief/015_intake.md (P1).
"""

from __future__ import annotations

import io
import os
from dataclasses import dataclass, field

# Firmas de archivo
_ZIP_MAGIC = b"PK\x03\x04"  # xlsx (es un zip)
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # xls legacy (compound file)
_UTF8_BOM = b"\xef\xbb\xbf"

# Cascada de encodings en modo estricto (orden = prioridad)
_ENCODING_CASCADE = ("utf-8", "cp1252", "latin-1")

# Delimitadores candidatos
_DELIMITADORES = (",", ";", "|", "\t")


@dataclass
class FormatSpec:
    """Resultado de la detección de formato de un snapshot.

    tipo              'csv' | 'xlsx' | 'xls' | None
    encoding          encoding CSV detectado (None en Excel)
    delimitador       ',' ';' '|' '\\t' (None en Excel)
    hoja              nombre de hoja Excel resuelta (None en CSV)
    fila_cabecera     fila de cabecera Excel, 1-based (None en CSV)
    hojas_disponibles lista de hojas cuando hay más de una (para escalar)
    ambiguo           True → la heurística no resolvió; escalar al operador
    corrupto          True → binario no tabular; P1 rechaza
    fuente            'detectado' | 'huella'
    """

    tipo: str | None = None
    encoding: str | None = None
    delimitador: str | None = None
    hoja: str | None = None
    fila_cabecera: int | None = None
    hojas_disponibles: list = field(default_factory=list)
    ambiguo: bool = False
    corrupto: bool = False
    fuente: str = "detectado"


def detect_format(raw_bytes: bytes, filename: str, huella: dict | None = None) -> FormatSpec:
    # 1) Huella de client_config (Excel ya confirmado en una entrega previa, etc.):
    #    respetar tal cual, sin re-detectar (brief P1).
    if huella:
        return FormatSpec(
            tipo=huella.get("tipo"),
            encoding=huella.get("encoding"),
            delimitador=huella.get("delimitador"),
            hoja=huella.get("hoja"),
            fila_cabecera=huella.get("fila_cabecera"),
            fuente="huella",
        )

    ext = os.path.splitext(filename)[1].lstrip(".").lower()

    # 2) Excel por magic bytes (la extensión por sí sola no es confiable).
    if raw_bytes[:4] == _ZIP_MAGIC:
        return _detect_xlsx(raw_bytes)
    if raw_bytes[:8] == _OLE2_MAGIC:
        return _detect_xls(raw_bytes)

    # 3) Camino texto/CSV. Un byte nulo delata un binario no tabular.
    if b"\x00" in raw_bytes:
        return FormatSpec(tipo=ext or None, corrupto=True)

    encoding, text = _detect_encoding(raw_bytes)
    if encoding is None:
        return FormatSpec(tipo=ext or None, corrupto=True)

    delimitador = _detect_delimiter(text)
    if delimitador is None:
        # texto decodificable pero sin estructura tabular reconocible → escalar
        return FormatSpec(tipo="csv", encoding=encoding, ambiguo=True)

    return FormatSpec(tipo="csv", encoding=encoding, delimitador=delimitador)


# --- Encoding (función pura, testeable de forma aislada) ---

def _detect_encoding(raw_bytes: bytes) -> tuple[str | None, str | None]:
    """Devuelve (encoding, texto) del primer encoding que decodifica en estricto.

    BOM UTF-8 se reconoce explícitamente como 'utf-8-sig'. latin-1 nunca falla
    (mapea los 256 bytes), por eso es el último recurso de la cascada.
    """
    if raw_bytes[:3] == _UTF8_BOM:
        try:
            return "utf-8-sig", raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            pass

    for enc in _ENCODING_CASCADE:
        try:
            return enc, raw_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    return None, None


# --- Delimitador (función pura) ---

def _detect_delimiter(text: str) -> str | None:
    """Elige el delimitador por consistencia de columnas entre filas.

    Prefiere el delimitador que produce el mismo nº de columnas en todas las
    filas de muestra y, a igualdad, el de más columnas. Si ninguno parte el
    texto en >1 columna, devuelve None (no es tabular).
    """
    lines = [ln for ln in text.splitlines() if ln.strip()][:20]
    if not lines:
        return None

    best = None
    best_score = -1
    for d in _DELIMITADORES:
        counts = [ln.count(d) for ln in lines]
        if max(counts) == 0:
            continue
        consistente = len(set(counts)) == 1
        ncols = counts[0] + 1
        score = (1000 if consistente else 0) + ncols
        if score > best_score:
            best_score = score
            best = d
    return best


# --- Excel ---

def _detect_xlsx(raw_bytes: bytes) -> FormatSpec:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception:
        return FormatSpec(tipo="xlsx", ambiguo=True)

    hojas = list(wb.sheetnames)
    if len(hojas) != 1:
        # múltiples hojas: no adivinar cuál — escalar al operador
        return FormatSpec(tipo="xlsx", hojas_disponibles=hojas,
                          hoja=hojas[0] if hojas else None, ambiguo=True)

    hoja = hojas[0]
    fila_cabecera = _primera_fila_con_datos(wb[hoja].iter_rows(values_only=True))
    wb.close()
    if fila_cabecera is None:
        return FormatSpec(tipo="xlsx", hoja=hoja, ambiguo=True)
    return FormatSpec(tipo="xlsx", hoja=hoja, fila_cabecera=fila_cabecera)


def _detect_xls(raw_bytes: bytes) -> FormatSpec:
    try:
        import xlrd

        book = xlrd.open_workbook(file_contents=raw_bytes)
    except Exception:
        # firma OLE2 confirma tipo 'xls', pero no se pudo resolver → escalar
        return FormatSpec(tipo="xls", ambiguo=True)

    hojas = book.sheet_names()
    if len(hojas) != 1:
        return FormatSpec(tipo="xls", hojas_disponibles=hojas,
                          hoja=hojas[0] if hojas else None, ambiguo=True)
    sh = book.sheet_by_index(0)
    fila_cabecera = None
    for i in range(sh.nrows):
        if any(str(v).strip() for v in sh.row_values(i)):
            fila_cabecera = i + 1
            break
    if fila_cabecera is None:
        return FormatSpec(tipo="xls", hoja=hojas[0], ambiguo=True)
    return FormatSpec(tipo="xls", hoja=hojas[0], fila_cabecera=fila_cabecera)


def _primera_fila_con_datos(rows) -> int | None:
    """Índice 1-based de la primera fila con al menos una celda no vacía."""
    for idx, fila in enumerate(rows, start=1):
        if any(c is not None and str(c).strip() != "" for c in fila):
            return idx
    return None
